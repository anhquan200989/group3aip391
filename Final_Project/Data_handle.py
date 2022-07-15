import sqlite3
import time

from openpyxl import load_workbook

# Connect to sqlite
conn = sqlite3.connect('CustomerInfo.sqlite')
#
# # Connect to cursor
cursor = conn.cursor()


# cursor.execute("drop table if exists BILL")
# cursor.execute("drop table if exists Item")
# cursor.execute("CREATE TABLE BILL ( ItemID INTEGER , Quantity INTEGER , Time TEXT , CustomerID INTEGER )")
# cursor.execute(
#     "CREATE TABLE Item (ID INTEGER , Name TEXT , Price FLOAT , Brand TEXT, Product_Line TEXT, PRIMARY KEY(ID) )")


# Clear table
def clear_Table(tableName):
    cursor.execute("delete from " + str(tableName))


# Add Item dataset from outside
def add_Item_from_file():
    lines = open('Item.txt').readlines()
    sql = "INSERT INTO Item VALUES(?,?,?,?,?)"
    num_line = 0
    for line in lines:
        num_line += 1
        data = line.rstrip().split()
        if (len(data) == 3):
            temp = data[2]
            data.pop(2)
            data.append("null")
            data.append(temp)
        if (num_line == len(lines)): break
        try:
            name = data[0]
            price = int(float(data[1]))
            brand = data[2]
            product_line = data[3]
            cursor.execute(sql, (num_line, name, price, brand, product_line))
        except Exception as e:
            print(e)
    # print("Add successfully")


def viewItem():
    cursor.execute("Select * from Item")
    catcher = cursor.fetchall()
    for item in catcher:
        print("{: <10} {: <30} {: <20} {: <20} {: <30}".format(*item))


def add_Item(name, price, brand, product_line):
    sql = "INSERT INTO Item VALUES(?,?,?,?,?)"
    id = cursor.execute("Select MAX(ID) from Item")
    try:
        file = open("Item.txt", "a")
        file.write(f'{name} {price} {brand} {product_line}')
        file.close()
        cursor.execute(sql, (id + 1, name, price, brand, product_line))
    except Exception as e:
        print(e)


def add_Bill():
    wb = load_workbook("BillData.xlsx")
    ws = wb.active

    Customer_id, item_id, quantity, time = getInput()
    sql = "INSERT INTO BILL VALUES(?,?,?,?)"
    for i in range(len(item_id)):
        try:
            cursor.execute(sql, (item_id[i], quantity[i], time, Customer_id))
            ws.append([item_id[i], quantity[i], time, Customer_id])
            wb.save("BillData.xlsx")
        except Exception as e:
            print(e)


def getInput():
    t = time.localtime()
    t = time.strftime('%Y-%m-%d %H:%M:%S', t)
    Customer_id = input('Enter customer id: ')
    item_id = []
    quantity = []
    print(f"Add all items in customer {Customer_id}'s bill:")
    while True:
        item_id.append(input('Enter item id: '))
        quantity.append(input('Enter quantity: '))
        choice = input('Is there any item left in this bill?(Y/N): ')
        if (choice in ['N', 'n']):
            print('Bill added')
            break
    return Customer_id, item_id, quantity, t


# Add Bill dataset from outside
def add_Bill_from_file():
    wb = load_workbook("BillData.xlsx")
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        try:
            ItemID = ws.cell(row=i, column=1).value
            Quantity = ws.cell(row=i, column=2).value
            Time = ws.cell(row=i, column=3).value
            CustomerID = ws.cell(row=i, column=4).value
            cursor.execute("INSERT INTO BILL VALUES(?,?,?,?)", (ItemID, Quantity, Time, CustomerID))
        except Exception as e:
            print(e)
    # print("Add bill Successfully\n")


def viewBill():
    cursor.execute("Select * from BILL")
    catcher = cursor.fetchall()
    if len(catcher) == 0:
        print("No order yet")
    else:
        for bill in catcher:
            print("{: <10} {: <20} {: <30} {: <10}".format(*bill))


def viewItem(ItemID):
    cursor.execute("Select * from Item where ID=?", (ItemID,))
    catcher = cursor.fetchall()
    return catcher[0][1]


def viewLatestOrder(CustomerID):
    # if customer's id is in db, but no order yet, bug
    cursor.execute("select * from BILL where CustomerID=? order by Time desc", (CustomerID,))
    catcher = cursor.fetchall()
    time = catcher[0][2]
    print(f"Latest bill by customer {CustomerID} in {time} ")
    for bill in catcher:
        if bill[2] == time:
            Item = viewItem(bill[0])
            Quan = bill[1]
            print(f"{Quan} {Item}")
    print('')


def isEmpty(tableName):
    cursor = conn.execute("SELECT count(*) FROM " + str(tableName))
    icount = cursor.fetchall()
    if (icount[0][0] == 0):
        return True
    else:
        return False
